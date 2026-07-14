import os
import json
from datetime import datetime, date

from flask import Blueprint, request, jsonify, current_app
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models.assessment_record import AssessmentRecord

assessment_bp = Blueprint("assessment_records", __name__)


def _allowed_file(filename):
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in current_app.config["ALLOWED_EXTENSIONS"]


def _parse_date(value):
    """Accepts 'YYYY-MM-DD' strings (what the frontend sends) or falls back to today."""
    if not value:
        return date.today()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except ValueError:
        return date.today()


def _parse_records_from_workbook(filepath):
    """
    Fallback server-side Excel parser, used only when the caller uploads a
    raw file without a pre-parsed `records` field (e.g. via Postman/curl
    instead of the React frontend, which parses with SheetJS in-browser).
    Only .xlsx is supported here; legacy .xls should be parsed client-side
    or converted first.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True)
    sheet = wb[wb.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    def find_col(*keys):
        for k in keys:
            if k in headers:
                return headers.index(k)
        return None

    idx_name = find_col("student_name", "student name", "name", "intern name")
    idx_assessment = find_col("assessment_name", "assessment name", "assessment", "title")
    idx_marks = find_col("marks", "score")
    idx_date = find_col("submitted_date", "submitted date", "date")

    records = []
    for row in rows[1:]:
        if idx_name is None or not row[idx_name]:
            continue
        raw_date = row[idx_date] if idx_date is not None else None
        submitted = raw_date.date().isoformat() if isinstance(raw_date, datetime) else (
            str(raw_date) if raw_date else date.today().isoformat()
        )
        records.append({
            "studentName": str(row[idx_name]).strip(),
            "assessmentName": str(row[idx_assessment]).strip() if idx_assessment is not None and row[idx_assessment] else "",
            "marks": float(row[idx_marks]) if idx_marks is not None and row[idx_marks] not in (None, "") else 0,
            "submittedDate": submitted,
        })
    return records


@assessment_bp.route("/upload", methods=["POST"])
def upload_assessment_records():
    """
    Bulk-create assessment records from an uploaded Excel file.

    Matches `uploadAssessmentRecords(file, parsedRecords)` in the frontend,
    which sends multipart/form-data with:
      - file: the raw .xlsx/.xls file (stored for audit purposes)
      - records: JSON string of already-parsed rows, each shaped like
        { studentName, assessmentName, marks, submittedDate }

    If `records` is omitted (e.g. calling this endpoint directly without
    the frontend's client-side SheetJS parsing), the server will attempt
    to parse the uploaded .xlsx file itself.
    """
    file = request.files.get("file")
    records_raw = request.form.get("records")

    if not file and not records_raw:
        return jsonify({"error": "No file or records provided"}), 400

    saved_filename = None
    filepath = None
    if file:
        if not _allowed_file(file.filename):
            return jsonify({"error": "Only .xlsx and .xls files are accepted"}), 400
        saved_filename = secure_filename(f"{datetime.utcnow().timestamp()}_{file.filename}")
        filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], saved_filename)
        file.save(filepath)

    if records_raw:
        try:
            parsed_records = json.loads(records_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "records must be a valid JSON array"}), 400
    else:
        try:
            parsed_records = _parse_records_from_workbook(filepath)
        except Exception:
            return jsonify({
                "error": "Could not parse the Excel file on the server. "
                         "Please supply pre-parsed 'records' JSON, or upload a .xlsx file."
            }), 400

    if not parsed_records:
        return jsonify({"error": "No valid records found to upload"}), 400

    created = []
    for r in parsed_records:
        student_name = str(r.get("studentName") or "").strip()
        if not student_name:
            continue

        record = AssessmentRecord(
            student_name=student_name,
            assessment_name=str(r.get("assessmentName") or "").strip(),
            marks=float(r.get("marks") or 0),
            submitted_date=_parse_date(r.get("submittedDate")),
            source_file=saved_filename,
        )
        db.session.add(record)
        created.append(record)

    db.session.commit()

    return jsonify({
        "message": f"{len(created)} record(s) uploaded",
        "records": [r.to_dict() for r in created],
    }), 201


@assessment_bp.route("", methods=["GET"])
def list_assessment_records():
    """
    List assessment records.
    Query params: search (matches student or assessment name), page, per_page
    """
    query = AssessmentRecord.query

    search = request.args.get("search", "").strip()
    if search:
        like = f"%{search}%"
        query = query.filter(
            db.or_(
                AssessmentRecord.student_name.ilike(like),
                AssessmentRecord.assessment_name.ilike(like),
            )
        )

    page = request.args.get("page", 1, type=int)
    per_page = min(request.args.get("per_page", 50, type=int), 200)

    query = query.order_by(AssessmentRecord.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        "records": [r.to_dict() for r in pagination.items],
        "total": pagination.total,
        "page": page,
        "per_page": per_page,
        "pages": pagination.pages,
    }), 200


@assessment_bp.route("/stats", methods=["GET"])
def assessment_stats():
    """Backs the three stat cards on the Assessment Management page."""
    records = AssessmentRecord.query.all()
    total = len(records)
    pass_mark = current_app.config["PASS_MARK"]

    pass_count = sum(1 for r in records if r.marks >= pass_mark)
    avg_marks = round(sum(r.marks for r in records) / total, 1) if total else 0

    return jsonify({
        "total": total,
        "passCount": pass_count,
        "avgMarks": avg_marks,
    }), 200


@assessment_bp.route("/<int:record_id>", methods=["GET"])
def get_assessment_record(record_id):
    record = AssessmentRecord.query.get(record_id)
    if not record:
        return jsonify({"error": "Record not found"}), 404
    return jsonify(record.to_dict()), 200


@assessment_bp.route("/<int:record_id>", methods=["PUT"])
def update_assessment_record(record_id):
    record = AssessmentRecord.query.get(record_id)
    if not record:
        return jsonify({"error": "Record not found"}), 404

    data = request.get_json(silent=True) or {}
    if "studentName" in data:
        record.student_name = str(data["studentName"]).strip()
    if "assessmentName" in data:
        record.assessment_name = str(data["assessmentName"]).strip()
    if "marks" in data:
        record.marks = float(data["marks"])
    if "submittedDate" in data:
        record.submitted_date = _parse_date(data["submittedDate"])

    db.session.commit()
    return jsonify(record.to_dict()), 200


@assessment_bp.route("/<int:record_id>", methods=["DELETE"])
def delete_assessment_record(record_id):
    """Matches `deleteAssessmentRecord(id)` in the frontend."""
    record = AssessmentRecord.query.get(record_id)
    if not record:
        return jsonify({"error": "Record not found"}), 404

    db.session.delete(record)
    db.session.commit()
    return jsonify({"message": "Record deleted"}), 200
